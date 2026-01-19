# -*- coding: utf-8 -*-
"""
Ferramentas Utilitárias - Excel, PDF, Gráficos, Web Search
"""

from typing import Dict, Any, List
from datetime import datetime
import logging
import os
import json
import base64

logger = logging.getLogger(__name__)


def ler_arquivo_excel(
    caminho: str,
    sheet: str = None,
    linhas: int = 10
) -> Dict[str, Any]:
    """
    Lê dados de um arquivo Excel.
    
    Args:
        caminho: Caminho do arquivo
        sheet: Nome da planilha (opcional)
        linhas: Número de linhas a retornar
        
    Returns:
        Dados do Excel
    """
    try:
        import openpyxl
        
        wb = openpyxl.load_workbook(caminho, read_only=True, data_only=True)
        
        # Selecionar sheet
        if sheet:
            ws = wb[sheet]
        else:
            ws = wb.active
        
        # Ler cabeçalhos
        headers = []
        for cell in ws[1]:
            headers.append(str(cell.value) if cell.value else f"Col{cell.column}")
        
        # Ler dados
        data = []
        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=linhas + 1), 1):
            row_data = {}
            for col_idx, cell in enumerate(row):
                row_data[headers[col_idx]] = cell.value
            data.append(row_data)
        
        wb.close()
        
        return {
            "arquivo": os.path.basename(caminho),
            "sheet": ws.title,
            "sheets_disponiveis": wb.sheetnames,
            "colunas": headers,
            "total_linhas": ws.max_row - 1,
            "linhas_retornadas": len(data),
            "dados": data,
            "status": "sucesso"
        }
        
    except ImportError:
        return {
            "erro": "Biblioteca openpyxl não instalada",
            "status": "erro"
        }
    except Exception as e:
        return {
            "erro": str(e),
            "status": "erro"
        }


def escrever_arquivo_excel(
    dados: List[Dict],
    nome_arquivo: str,
    sheet: str = "Dados"
) -> Dict[str, Any]:
    """
    Escreve dados em um arquivo Excel.
    
    Args:
        dados: Lista de dicionários com dados
        nome_arquivo: Nome do arquivo de saída
        sheet: Nome da planilha
        
    Returns:
        Informações do arquivo criado
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill
        
        # Garantir diretório de artefatos
        from ..config import get_config
        config = get_config()
        artifacts_path = config.artifacts_path
        os.makedirs(artifacts_path, exist_ok=True)
        
        caminho = os.path.join(artifacts_path, nome_arquivo)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet
        
        if not dados:
            return {"erro": "Dados vazios", "status": "erro"}
        
        # Cabeçalhos
        headers = list(dados[0].keys())
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        
        # Dados
        for row_idx, row_data in enumerate(dados, 2):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(header))
        
        wb.save(caminho)
        
        return {
            "arquivo": nome_arquivo,
            "caminho": caminho,
            "sheet": sheet,
            "linhas": len(dados),
            "colunas": len(headers),
            "status": "sucesso"
        }
        
    except Exception as e:
        return {"erro": str(e), "status": "erro"}


def ler_arquivo_pdf(
    caminho: str,
    paginas: int = 5
) -> Dict[str, Any]:
    """
    Extrai texto de um arquivo PDF.
    
    Args:
        caminho: Caminho do arquivo
        paginas: Número de páginas a ler
        
    Returns:
        Texto extraído
    """
    try:
        import pypdf
        
        reader = pypdf.PdfReader(caminho)
        total_pages = len(reader.pages)
        
        texto = []
        for i in range(min(paginas, total_pages)):
            page = reader.pages[i]
            texto.append({
                "pagina": i + 1,
                "texto": page.extract_text()
            })
        
        texto_completo = "\n\n".join([p["texto"] for p in texto])
        
        return {
            "arquivo": os.path.basename(caminho),
            "total_paginas": total_pages,
            "paginas_lidas": len(texto),
            "conteudo": texto,
            "texto_completo": texto_completo[:5000],  # Limitar tamanho
            "status": "sucesso"
        }
        
    except ImportError:
        return {"erro": "Biblioteca pypdf não instalada", "status": "erro"}
    except Exception as e:
        return {"erro": str(e), "status": "erro"}


def gerar_grafico(
    tipo: str,
    dados: List[Dict],
    titulo: str = "Gráfico",
    eixo_x: str = None,
    eixo_y: str = None
) -> Dict[str, Any]:
    """
    Gera um gráfico usando matplotlib.
    
    Args:
        tipo: Tipo do gráfico (bar, line, pie, scatter)
        dados: Dados para o gráfico
        titulo: Título do gráfico
        eixo_x: Campo para eixo X
        eixo_y: Campo para eixo Y
        
    Returns:
        Gráfico em base64
    """
    try:
        import matplotlib
        matplotlib.use('Agg')
        import matplotlib.pyplot as plt
        import io
        
        fig, ax = plt.subplots(figsize=(10, 6))
        
        if not dados:
            return {"erro": "Dados vazios", "status": "erro"}
        
        # Determinar eixos automaticamente se não fornecidos
        if not eixo_x or not eixo_y:
            keys = list(dados[0].keys())
            eixo_x = eixo_x or keys[0]
            eixo_y = eixo_y or (keys[1] if len(keys) > 1 else keys[0])
        
        x_vals = [d.get(eixo_x, i) for i, d in enumerate(dados)]
        y_vals = [d.get(eixo_y, 0) for d in dados]
        
        if tipo == "bar":
            ax.bar(range(len(x_vals)), y_vals, color='#366092')
            ax.set_xticks(range(len(x_vals)))
            ax.set_xticklabels(x_vals, rotation=45, ha='right')
        elif tipo == "line":
            ax.plot(x_vals, y_vals, marker='o', color='#366092')
        elif tipo == "pie":
            ax.pie(y_vals, labels=x_vals, autopct='%1.1f%%')
        elif tipo == "scatter":
            ax.scatter(x_vals, y_vals, color='#366092')
        else:
            ax.bar(range(len(x_vals)), y_vals)
        
        ax.set_title(titulo, fontsize=14, fontweight='bold')
        if tipo != "pie":
            ax.set_xlabel(eixo_x)
            ax.set_ylabel(eixo_y)
        
        plt.tight_layout()
        
        # Salvar em memória e converter para base64
        buffer = io.BytesIO()
        plt.savefig(buffer, format='png', dpi=100)
        buffer.seek(0)
        img_base64 = base64.b64encode(buffer.read()).decode()
        plt.close(fig)
        
        # Salvar arquivo também
        from ..config import get_config
        config = get_config()
        artifacts_path = config.artifacts_path
        os.makedirs(artifacts_path, exist_ok=True)
        
        nome_arquivo = f"grafico_{tipo}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.png"
        caminho = os.path.join(artifacts_path, nome_arquivo)
        
        with open(caminho, 'wb') as f:
            buffer.seek(0)
            f.write(buffer.read())
        
        return {
            "tipo": tipo,
            "titulo": titulo,
            "arquivo": nome_arquivo,
            "caminho": caminho,
            "imagem_base64": f"data:image/png;base64,{img_base64[:100]}...",  # Truncado
            "pontos_dados": len(dados),
            "status": "sucesso"
        }
        
    except ImportError:
        return {"erro": "Biblioteca matplotlib não instalada", "status": "erro"}
    except Exception as e:
        return {"erro": str(e), "status": "erro"}


async def pesquisar_web(
    query: str,
    num_resultados: int = 5
) -> Dict[str, Any]:
    """
    Realiza pesquisa na web.
    
    Args:
        query: Termo de pesquisa
        num_resultados: Número de resultados
        
    Returns:
        Resultados da pesquisa
    """
    try:
        import httpx
        
        # Usar DuckDuckGo (não requer API key)
        async with httpx.AsyncClient() as client:
            response = await client.get(
                "https://api.duckduckgo.com/",
                params={
                    "q": query,
                    "format": "json",
                    "no_html": 1,
                    "skip_disambig": 1
                },
                timeout=10.0
            )
            
            if response.status_code == 200:
                data = response.json()
                
                resultados = []
                
                # Abstract
                if data.get("Abstract"):
                    resultados.append({
                        "titulo": data.get("Heading", "Resumo"),
                        "descricao": data["Abstract"],
                        "url": data.get("AbstractURL", ""),
                        "fonte": data.get("AbstractSource", "")
                    })
                
                # Related Topics
                for topic in data.get("RelatedTopics", [])[:num_resultados]:
                    if isinstance(topic, dict) and "Text" in topic:
                        resultados.append({
                            "titulo": topic.get("Text", "")[:100],
                            "descricao": topic.get("Text", ""),
                            "url": topic.get("FirstURL", "")
                        })
                
                return {
                    "query": query,
                    "total_resultados": len(resultados),
                    "resultados": resultados,
                    "status": "sucesso"
                }
            
            return {"erro": f"Erro na pesquisa: {response.status_code}", "status": "erro"}
            
    except Exception as e:
        return {"erro": str(e), "status": "erro"}


# Schema das ferramentas
UTILS_TOOLS = [
    {
        "name": "ler_arquivo_excel",
        "description": "Lê dados de um arquivo Excel (XLSX). Retorna cabeçalhos, dados e informações do arquivo.",
        "parameters": {
            "type": "object",
            "properties": {
                "caminho": {"type": "string", "description": "Caminho completo do arquivo"},
                "sheet": {"type": "string", "description": "Nome da planilha (opcional)"},
                "linhas": {"type": "integer", "description": "Número de linhas a ler (padrão 10)"}
            },
            "required": ["caminho"]
        }
    },
    {
        "name": "escrever_arquivo_excel",
        "description": "Cria um arquivo Excel com os dados fornecidos.",
        "parameters": {
            "type": "object",
            "properties": {
                "dados": {"type": "array", "description": "Lista de objetos com dados", "items": {"type": "object"}},
                "nome_arquivo": {"type": "string", "description": "Nome do arquivo de saída (com .xlsx)"},
                "sheet": {"type": "string", "description": "Nome da planilha (padrão 'Dados')"}
            },
            "required": ["dados", "nome_arquivo"]
        }
    },
    {
        "name": "ler_arquivo_pdf",
        "description": "Extrai texto de um arquivo PDF.",
        "parameters": {
            "type": "object",
            "properties": {
                "caminho": {"type": "string", "description": "Caminho completo do arquivo PDF"},
                "paginas": {"type": "integer", "description": "Número de páginas a ler (padrão 5)"}
            },
            "required": ["caminho"]
        }
    },
    {
        "name": "gerar_grafico",
        "description": "Gera um gráfico (bar, line, pie, scatter) a partir dos dados fornecidos.",
        "parameters": {
            "type": "object",
            "properties": {
                "tipo": {"type": "string", "description": "Tipo: bar, line, pie, scatter"},
                "dados": {"type": "array", "description": "Lista de objetos com dados", "items": {"type": "object"}},
                "titulo": {"type": "string", "description": "Título do gráfico"},
                "eixo_x": {"type": "string", "description": "Campo para eixo X"},
                "eixo_y": {"type": "string", "description": "Campo para eixo Y"}
            },
            "required": ["tipo", "dados"]
        }
    },
    {
        "name": "pesquisar_web",
        "description": "Realiza pesquisa na web e retorna resultados resumidos.",
        "parameters": {
            "type": "object",
            "properties": {
                "query": {"type": "string", "description": "Termo de pesquisa"},
                "num_resultados": {"type": "integer", "description": "Número de resultados (padrão 5)"}
            },
            "required": ["query"]
        }
    }
]


__all__ = [
    "ler_arquivo_excel",
    "escrever_arquivo_excel",
    "ler_arquivo_pdf",
    "gerar_grafico",
    "pesquisar_web",
    "UTILS_TOOLS"
]
